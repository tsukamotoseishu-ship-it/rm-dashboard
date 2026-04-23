"""
RM価格案 Excel出力スクリプト - ホテル甲子園
インプットデータは2パターン自動判定:
  [優先] PMSデータ (a.csv) - 精緻・直予約含む・部屋名確定
  [代替] ラクツウCSV (ReserveList_*.csv) - PMSがない場合のフォールバック

実行方法: python export_rm_excel.py
"""

import csv, glob, math, os
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 設定
# ============================================================
OUTPUT_PATH      = r"C:\Users\tsukamoto.seishu\rm_system\RM価格案.xlsx"
CSV_DIR          = r"C:\Users\tsukamoto.seishu\Downloads"
PMS_CSV          = r"C:\Users\tsukamoto.seishu\Downloads\a.csv"   # PMSデータ（優先）
COMP_PRICES_CSV  = r"C:\Users\tsukamoto.seishu\rm_system\competitor_prices.csv"
TODAY            = datetime(2026, 4, 9)
DAYS_AHEAD       = 30
TOTAL_ROOMS      = 17   # 17室（天神・天目・昇仙峡・夜叉神・本栖・笛吹・金峰・赤岳・風林
                         #       ・白糸・千鳥・甲斐駒・白凰・釜無・精進・薬師・御坂）

# 月次サマリー用：月次予算（YYYYMM → 金額）
# ※スプレッドシートから転記。変更はここを編集してください。
MONTHLY_BUDGET = {
    '202404': 13_000_000,
    '202405': 14_000_000,
    '202406': 12_000_000,
    '202407': 18_000_000,
    '202408': 20_000_000,
    '202409': 16_000_000,
    '202410': 15_000_000,
    '202411': 13_000_000,
    '202412': 18_447_404,
    '202501': 17_200_984,
    '202502': 13_653_373,
    '202503': 18_618_641,
    '202504': 17_419_169,
    '202505': 15_498_589,
    '202506': 13_605_922,
    '202507': 18_000_000,
    '202508': 20_000_000,
    '202509': 16_000_000,
}

RANKS = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1',
         'K1','L1','M1','N1','O1','P1','Q1','R1']
RANK_PRICE = {
    'A1':13300,'B1':15300,'C1':16300,'D1':17800,'E1':19300,'F1':20800,
    'G1':22300,'H1':23800,'I1':25300,'J1':26800,'K1':28300,'L1':29800,
    'M1':31300,'N1':32800,'O1':34300,'P1':35800,'Q1':37300,'R1':38800,
}
ROOM_NAMES = ['金峰','笛吹','天神','天目','赤岳','風林']
ROOM_RANKS_BASE = {
    '金峰': ['E1','F1','G1','M1','F1','D1','B1','C1','E1','F1','G1','F1','E1','D1'],
    '笛吹': ['F1','G1','G1','L1','F1','E1','C1','D1','F1','G1','H1','G1','F1','E1'],
    '天神': ['E1','F1','G1','M1','F1','D1','B1','C1','E1','F1','G1','F1','E1','D1'],
    '天目': ['K1','L1','M1','R1','L1','K1','H1','I1','K1','L1','M1','L1','K1','J1'],
    '赤岳': ['D1','E1','F1','L1','E1','D1','B1','C1','D1','E1','F1','E1','D1','C1'],
    '風林': ['F1','G1','H1','N1','G1','E1','C1','D1','F1','G1','H1','G1','F1','E1'],
}

# competitor_prices.csv の施設名 → 短縮名マッピング
COMP_NAME_MAP = {
    '【2食】石和温泉　銘石の宿　かげつ':         'かげつ',
    '【2食】笛吹川温泉　坐忘':                   '坐忘',
    '【2食】石和温泉　ホテルふじ':               'ホテルふじ',
    '【2食】石和温泉　華やぎの章　慶山':         '慶山',
    '【2食】石和温泉　ホテル古柏園':             '古柏園',
    '【2食】石和温泉　糸柳こやど　ゆわ':         '糸柳こやどゆわ',
    '【2食】川浦温泉　山県館':                   '山県館',
    '【2食】糸柳別館　離れの邸　和穣苑':         '糸柳別館',
    '【2食】別邸　花水晶':                       '花水晶',
    '【2食】石和名湯館　糸柳':                   '糸柳',
    '【2食】湯めぐり宿　笛吹川':                 '笛吹川',
    '【2食】石和温泉郷　旅館深雪温泉':           '深雪温泉',
    '【2食】銘庭の宿　ホテル甲子園':             'ホテル甲子園（自社）',
}

# ============================================================
# 色・スタイルヘルパー
# ============================================================
def argb(h): return 'FF' + h.lstrip('#')

C = {
    'dark':   argb('#1a3a5c'), 'blue':   argb('#2980b9'), 'white': argb('#FFFFFF'),
    'gray':   argb('#F8F9FA'), 'lgray':  argb('#EEEEEE'), 'sub':   argb('#EBF5FB'),
    'sat':    argb('#FFF9E6'), 'hol':    argb('#FFF0F0'), 'altrow':argb('#FAFAFA'),
    'up_bg':  argb('#EBF5FB'), 'dn_bg':  argb('#FDEDEC'), 'ok_bg': argb('#EAFAF1'),
    'red':    argb('#E74C3C'), 'green':  argb('#27AE60'), 'orange':argb('#E67E22'),
    'purple': argb('#8E44AD'), 'yellow': argb('#F9CA24'),
}

def fill(c): return PatternFill('solid', fgColor=c)
def font(sz=10, bold=False, color='FF333333', name='Meiryo UI'):
    return Font(name=name, size=sz, bold=bold, color=color)
def side(style='thin', color='FFD0D0D0'): return Side(style=style, color=color)
def border(): return Border(left=side(), right=side(), top=side(), bottom=side())
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

def style(ws, row, col, value=None, sz=10, bold=False, fg='FF333333',
          bg=None, align='center', num_format=None, border_on=True):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    c.font = font(sz, bold, fg)
    if bg: c.fill = fill(bg)
    c.alignment = center() if align == 'center' else left()
    if border_on: c.border = border()
    if num_format: c.number_format = num_format
    return c

# ============================================================
# データ読み込み
# ============================================================
def load_data():
    """
    インプットデータを2パターン自動判定して読み込む。
    [優先] PMS CSV (a.csv) : 精緻・直予約含む・実際の部屋名
    [代替] ラクツウ CSV (ReserveList_*.csv) : PMSがない場合のフォールバック
    戻り値: (daily, lead_dist, comp_prices, data_source_label)
    """

    # ======================================================
    # PMS CSV 読み込み関数
    # ======================================================
    def load_pms(fpath):
        """
        a.csv 形式のPMSデータを読み込む。
        - 利用有無=有 かつ 科目に「宿泊」を含む行のみ対象
        - 予約番号+宿泊部屋 でユニーク集計（複数行ある予約を1室として扱う）
        - 利用日+泊数 から各泊の日付を展開
        """
        with open(fpath, encoding='cp932') as f:
            raw = list(csv.DictReader(f))

        # 宿泊行を抽出（キャンセル除外）
        stay_rows = [r for r in raw
                     if r.get('利用有無','') == '有'
                     and '宿泊' in r.get('科目','')]

        # 日別ユニーク室セット（予約番号_部屋名 で重複排除）
        daily_rooms = defaultdict(set)
        lead_list   = []   # (利用日, 予約日) のリスト（ブッキングカーブ用）
        seen_res    = set() # ブッキングカーブ用重複排除

        for r in stay_rows:
            try:
                stay_date = datetime.strptime(r['利用日'], '%Y%m%d')
                nights    = int(r.get('泊数', 1) or 1)
                room      = r['宿泊部屋'].strip().lstrip('*')  # *付きフラグを除去して正規化
                res_no    = r['予約番号'].strip()
                if not room or ',' in room: continue  # 複数部屋割当は除外

                # 各泊の日付に展開（部屋名でユニーク化 = 同室二重予約を1室として数える）
                for n in range(nights):
                    d = stay_date + timedelta(days=n)
                    daily_rooms[d].add(room)

                # ブッキングカーブ用（予約番号単位で1回だけ）
                res_key = f'{res_no}_{stay_date.strftime("%Y%m%d")}'
                if res_key not in seen_res:
                    booked = datetime.strptime(r['予約日'], '%Y%m%d')
                    ld = (stay_date - booked).days
                    if 0 <= ld <= 180:
                        lead_list.append((stay_date, ld))
                    seen_res.add(res_key)
            except: pass

        daily = {d: len(s) for d, s in daily_rooms.items()}

        lead_dist = defaultdict(lambda: defaultdict(int))
        for stay_date, ld in lead_list:
            wd = stay_date.weekday()
            dt = '土/連休' if wd == 5 else ('日曜' if wd == 6 else '平日')
            lead_dist[dt][ld // 7] += 1

        # 月別売上集計（全科目・利用有無=有の金額を合算）
        monthly_rev = defaultdict(float)
        for r in raw:
            if r.get('利用有無', '') == '有':
                try:
                    stay_date = datetime.strptime(r['利用日'], '%Y%m%d')
                    month_key = stay_date.strftime('%Y%m')
                    amount = float(r.get('金額', '0') or 0)
                    monthly_rev[month_key] += amount
                except:
                    pass

        return daily, lead_dist, dict(monthly_rev)

    # ======================================================
    # ラクツウ CSV 読み込み関数
    # ======================================================
    def load_rakutsuu():
        files = glob.glob(f'{CSV_DIR}/ReserveList_*.csv')
        rows = {}
        for fp in sorted(files):
            with open(fp, encoding='cp932') as f:
                for r in csv.DictReader(f):
                    k = r.get('予約番号','')
                    if k and k not in rows and r.get('区分','') != 'キャンセル':
                        rows[k] = r

        daily = defaultdict(int)
        for r in rows.values():
            try:
                cin  = datetime.strptime(r['チェックイン'], '%Y%m%d')
                cout = datetime.strptime(r['チェックアウト'], '%Y%m%d')
                rms  = int(r.get('室数', 1))
                d = cin
                while d < cout:
                    daily[d] += rms
                    d += timedelta(days=1)
            except: pass

        lead_dist = defaultdict(lambda: defaultdict(int))
        for r in rows.values():
            try:
                cin = datetime.strptime(r['チェックイン'], '%Y%m%d')
                reg = datetime.strptime(r['受信日／登録日'][:8], '%Y%m%d')
                ld  = (cin - reg).days
                if 0 <= ld <= 180:
                    wd = cin.weekday()
                    dt = '土/連休' if wd == 5 else ('日曜' if wd == 6 else '平日')
                    lead_dist[dt][ld // 7] += 1
            except: pass

        return dict(daily), lead_dist, {}  # ラクツウは金額情報なし

    # ======================================================
    # データソース選択
    # ======================================================
    if os.path.exists(PMS_CSV):
        daily, lead_dist, monthly_rev = load_pms(PMS_CSV)
        data_source = f'PMSデータ ({os.path.basename(PMS_CSV)})'
    else:
        daily, lead_dist, monthly_rev = load_rakutsuu()
        data_source = 'ラクツウCSV (ReserveList_*.csv)'

    # ======================================================
    # 競合価格CSV
    # ======================================================
    comp_prices = defaultdict(dict)
    try:
        with open(COMP_PRICES_CSV, encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                name      = r.get('施設名','').strip()
                short     = COMP_NAME_MAP.get(name, name)
                date      = r.get('対象日','').strip()
                price_str = r.get('最低価格','×').strip()
                price = None
                if price_str != '×':
                    try: price = int(price_str)
                    except: pass
                comp_prices[date][short] = price
    except FileNotFoundError:
        pass

    return daily, lead_dist, comp_prices, data_source, monthly_rev

# ============================================================
# モック/計算ヘルパー
# ============================================================
def day_type(d):
    wd = d.weekday()
    hols = {(1,1),(1,2),(1,3),(4,29),(5,3),(5,4),(5,5),(7,21),(8,11),(9,15),(10,13),(11,3),(11,23),(12,23)}
    if wd == 5 or (d.month, d.day) in hols: return '土/連休初日'
    if wd == 6: return '平日/日'
    return '平日/日'


def comp_price(hidx, day_off):
    bases = [48000,68000,35000,40000,31000,55000,37000,29000,73000]
    b = bases[hidx % len(bases)]
    wk = 1.3 if day_off % 7 in (5,6) else 1.0
    n  = 1 + math.sin(hidx*3.1 + day_off*0.7)*0.15
    return round(b * wk * n / 100) * 100

def get_rank(room, idx):
    ranks = ROOM_RANKS_BASE[room]
    return ranks[min(idx, len(ranks)-1)]

def suggest_rank(cur, action):
    i = RANKS.index(cur)
    if action == 'UP'   and i < len(RANKS)-1: return RANKS[i+1]
    if action == 'DOWN' and i > 0:            return RANKS[i-1]
    return cur

WDAYS = ['月','火','水','木','金','土','日']

# ============================================================
# ブッキングカーブ（実データから算出: N日前時点で最終予約の何%が入っているか）
# 土曜・日曜・平日でほぼ同じ傾向のため共通カーブを使用
# ============================================================
BOOKING_CURVE = {
    # (lead_days, cumulative_pct)
    0: 1.000, 1: 0.953, 2: 0.938, 3: 0.826, 4: 0.812,
    5: 0.764, 6: 0.750, 7: 0.704, 8: 0.690, 9: 0.676,
    10: 0.637, 11: 0.625, 12: 0.613, 13: 0.601, 14: 0.540,
    15: 0.530, 16: 0.520, 17: 0.510, 18: 0.500, 19: 0.490,
    20: 0.480, 21: 0.407, 25: 0.360, 30: 0.270, 45: 0.138,
    60: 0.071, 90: 0.020,
}
_CURVE_KEYS = sorted(BOOKING_CURVE.keys())

def booking_curve_at(lead_days):
    """N日前時点での進捗率（線形補間）"""
    ld = max(0, lead_days)
    if ld >= _CURVE_KEYS[-1]:
        return BOOKING_CURVE[_CURVE_KEYS[-1]]
    for i, k in enumerate(_CURVE_KEYS):
        if k >= ld:
            if k == ld or i == 0:
                return BOOKING_CURVE[k]
            k0, k1 = _CURVE_KEYS[i-1], k
            v0, v1 = BOOKING_CURVE[k0], BOOKING_CURVE[k1]
            return v0 + (v1 - v0) * (ld - k0) / (k1 - k0)
    return BOOKING_CURVE[_CURVE_KEYS[-1]]

# 目標最終稼働率（17室ベースの実績中央値ベース）
# 実績（17室）: 土曜88%, 金曜71%, 平日55%, 日曜71%
# 目標: 実績中央値を基準に設定
TARGET_FINAL_OCC = {
    '土/連休初日': 0.85,   # 実績中央値88%、目標は少し保守的に
    '平日/日':     0.65,   # 金・日の中央値70%、平日55% → 平均65%
}

def target_occ(lead, dtype):
    """ブッキングカーブを考慮した『今この時点での期待進捗率』"""
    final = TARGET_FINAL_OCC.get(dtype, 0.45)
    curve = booking_curve_at(lead)
    return final * curve

# ============================================================
# ヘッダー行ユーティリティ
# ============================================================
def write_header_row(ws, row_num, height, cols):
    """cols: list of (col_idx, value, bg, fg, sz, bold, merge_to_col)"""
    ws.row_dimensions[row_num].height = height
    for item in cols:
        col, val, bg, fg, sz, bold, merge_end = item
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = font(sz, bold, fg)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border()
        if merge_end:
            ws.merge_cells(
                start_row=row_num, start_column=col,
                end_row=row_num, end_column=merge_end
            )

def title_row(ws, row_num, text, ncols, height=26):
    ws.row_dimensions[row_num].height = height
    c = ws.cell(row=row_num, column=1, value=text)
    c.font = font(12, True, C['white'])
    c.fill = fill(C['dark'])
    c.alignment = center()
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=ncols)

def note_row(ws, row_num, text, ncols, height=16):
    ws.row_dimensions[row_num].height = height
    c = ws.cell(row=row_num, column=1, value=text)
    c.font = font(9, False, argb('#666666'), italic=True) if False else font(9, False, argb('#666666'))
    c.fill = fill(C['gray'])
    c.alignment = left()
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=ncols)

# ============================================================
# シート1: RM価格案（メイン判断表）
# ============================================================
def comp_avg_for_date(comp_prices, date_str, exclude='ホテル甲子園（自社）'):
    """指定日の競合平均価格（自社除く、価格ありのみ平均）"""
    day_data = comp_prices.get(date_str, {})
    prices = [v for k, v in day_data.items() if k != exclude and v is not None]
    return round(sum(prices) / len(prices)) if prices else None

def build_rm_sheet(wb, daily, comp_prices, data_source=''):
    ws = wb.create_sheet("RM価格案")
    ws.sheet_view.showGridLines = False

    NCOLS = 15
    title_row(ws, 1,
        f"石和温泉 ホテル甲子園　RM価格案　取得日: {TODAY.strftime('%Y/%m/%d')}",
        NCOLS)
    src_label = f'予約データ: {data_source}' if data_source else ''
    note_row(ws, 2,
        f"【判断ロジック】ブッキングカーブ進捗率×目標最終稼働率(土85%/平日65%)=期待進捗、実予約がこれを上回ればUP・下回ればDOWN　／　{src_label}",
        NCOLS)

    # ヘッダー
    hdrs = ['宿泊日','曜日','リードタイム','日判定',
            '期待進捗率\n(目標×カーブ)',
            '実績\n消化率',
            '差異\n(実-期待)',
            '競合\n平均価格\n(2名合計)','価格\n差異','現在ランク\n(金峰代表)',
            '現在価格\n(2名合計)','推奨\nアクション','推奨ランク','推奨価格\n(2名合計)','備考']
    ws.row_dimensions[3].height = 36
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = font(9, True, C['white'])
        c.fill = fill(C['dark'])
        c.alignment = center()
        c.border = border()

    widths = [11,5,10,12,8,14,7,14,9,13,11,10,10,10,22]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 4
    rows_data = []
    for i in range(1, DAYS_AHEAD+1):
        d = TODAY + timedelta(days=i)
        dtype = day_type(d)
        lead  = i

        target      = target_occ(lead, dtype)
        actual_note = ''

        if d > TODAY:
            # 将来日程: 今回の予約CSVの実数値（現時点での予約済み室数）
            actual = daily.get(d, 0) / TOTAL_ROOMS
            if daily.get(d, 0) == 0:
                actual_note = '予約なし'
        else:
            # 過去日程: 前年同週の実績
            d_ly   = d - timedelta(days=364)
            actual = daily.get(d_ly, 0) / TOTAL_ROOMS

        diff = actual - target

        # 競合価格（実データ優先、なければモック）
        date_str = d.strftime('%Y/%m/%d')
        cavg_real = comp_avg_for_date(comp_prices, date_str)
        if cavg_real is not None:
            cavg = cavg_real
            comp_note = ''
        else:
            cp = [comp_price(hi, i) for hi in range(9)]
            cavg = round(sum(cp)/len(cp))
            comp_note = '(モック)'
        ri       = min(i-1, len(ROOM_RANKS_BASE['金峰'])-1)
        cur_rank = ROOM_RANKS_BASE['金峰'][ri]
        cur_p    = RANK_PRICE.get(cur_rank, 0) * 2   # 2名合計

        own_p = cur_p   # 自社価格（2名合計）と競合を比較
        pdiff = own_p - cavg

        if diff >= 0.05:    action = 'UP'
        elif diff <= -0.10: action = 'DOWN'
        else:               action = 'STAY'

        sug_rank = suggest_rank(cur_rank, action)
        sug_p    = RANK_PRICE.get(sug_rank, 0) * 2   # 2名合計

        note = ''
        if dtype == '土/連休初日': note = '土曜・連休'
        if cavg > 50000:           note += (' ' if note else '') + '競合高値帯'
        if diff < -0.20:           note += (' ' if note else '') + '消化率注意'
        if comp_note:              note += (' ' if note else '') + comp_note
        if actual_note:            note += (' ' if note else '') + actual_note

        wd = WDAYS[d.weekday()]
        if dtype == '土/連休初日':   row_bg = C['sat']
        elif d.weekday() == 6:       row_bg = C['hol']
        else:                        row_bg = C['white'] if i%2==0 else C['altrow']

        vals = [d.strftime('%m/%d'), wd, f'{lead}日前', dtype,
                target, actual, diff,
                cavg, pdiff, cur_rank, cur_p,
                action, sug_rank, sug_p, note]

        ws.row_dimensions[row].height = 20
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill(row_bg)
            c.border = border()
            c.alignment = center()
            c.font = font(10)

            if ci == 1:  c.font = font(10, True)
            elif ci == 2:
                if val == '土': c.font = font(10, True, C['orange'])
                elif val == '日': c.font = font(10, True, C['red'])
            elif ci in (5,6): c.number_format = '0.0%'
            elif ci == 7:
                c.number_format = '+0.0%;-0.0%;"-"'
                c.font = font(10, True, C['green'] if diff >= 0.05 else
                              (C['red'] if diff <= -0.10 else argb('#888888')))
            elif ci in (8,11,14): c.number_format = '#,##0"円"'
            elif ci == 9:
                c.number_format = '+#,##0"円";-#,##0"円";"-"'
                c.font = font(10, True, C['green'] if pdiff>=0 else C['red'])
            elif ci == 10: c.font = font(10, True, C['dark'])
            elif ci == 12:
                bg = {'UP':C['up_bg'],'DOWN':C['dn_bg'],'STAY':C['ok_bg']}[action]
                fg = {'UP':C['blue'],'DOWN':C['red'],'STAY':C['green']}[action]
                c.fill = fill(bg); c.font = font(10, True, fg)
            elif ci == 13:
                changed = sug_rank != cur_rank
                c.font = font(10, True, C['blue'] if action=='UP' else
                              C['red'] if action=='DOWN' else argb('#AAAAAA'))
                if changed: c.fill = fill(C['up_bg'] if action=='UP' else C['dn_bg'])
            elif ci == 15:
                c.alignment = left(); c.font = font(9, False, argb('#888888'))

        rows_data.append(dict(d=d, dtype=dtype, target=target, actual=actual,
                               diff=diff, cavg=cavg, cur_rank=cur_rank, action=action))
        row += 1

    ws.freeze_panes = 'A4'
    return ws, rows_data

# ============================================================
# シート2: 部屋別ランク設定
# ============================================================
def build_rank_sheet(wb, rows_data):
    ws = wb.create_sheet("部屋別ランク設定")
    ws.sheet_view.showGridLines = False
    NCOLS = 2 + DAYS_AHEAD

    title_row(ws, 1, f"部屋タイプ別 推奨ランク設定表（モック）　{TODAY.strftime('%Y/%m/%d')} 時点", NCOLS)
    note_row(ws, 2, "■ 現在ランク（上段）　■ 推奨ランク（下段）：青=UP余地、赤=DOWN推奨", NCOLS)

    # 列ヘッダー
    for ci, label in enumerate(['部屋タイプ','リードタイム'], 1):
        ws.merge_cells(start_row=3, start_column=ci, end_row=4, end_column=ci)
        c = ws.cell(row=3, column=ci, value=label)
        c.font = font(9, True, C['white']); c.fill = fill(C['dark'])
        c.alignment = center(); c.border = border()
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10

    for i in range(1, DAYS_AHEAD+1):
        d = TODAY + timedelta(days=i)
        dtype = day_type(d)
        is_sat = dtype == '土/連休初日'
        col = i+2
        bg = C['orange'] if is_sat else C['blue']

        ws.row_dimensions[3].height = 18; ws.row_dimensions[4].height = 16
        c3 = ws.cell(row=3, column=col, value=f"{d.month}/{d.day}")
        c3.font = font(9, True, C['white']); c3.fill = fill(bg)
        c3.alignment = center(); c3.border = border()
        c4 = ws.cell(row=4, column=col, value=WDAYS[d.weekday()])
        c4.font = font(9, True, C['white']); c4.fill = fill(bg)
        c4.alignment = center(); c4.border = border()
        ws.column_dimensions[get_column_letter(col)].width = 6.5

    for ri, room in enumerate(ROOM_NAMES):
        r_cur = 5 + ri*3
        r_sug = r_cur + 1
        r_spc = r_cur + 2
        ws.row_dimensions[r_cur].height = 18
        ws.row_dimensions[r_sug].height = 18
        ws.row_dimensions[r_spc].height = 5

        ws.merge_cells(start_row=r_cur, start_column=1, end_row=r_sug, end_column=1)
        c = ws.cell(row=r_cur, column=1, value=room)
        c.font = font(11, True, C['dark']); c.fill = fill(C['sub'])
        c.alignment = center(); c.border = border()

        for lbl, row_n in [('現在ランク', r_cur), ('推奨ランク', r_sug)]:
            c = ws.cell(row=row_n, column=2, value=lbl)
            c.font = font(9, False, argb('#666666')); c.fill = fill(C['gray'])
            c.alignment = center(); c.border = border()

        for i in range(1, DAYS_AHEAD+1):
            row_d = rows_data[i-1] if i-1 < len(rows_data) else {}
            action   = row_d.get('action', 'STAY')
            cur_rank = get_rank(room, i-1)
            sug_rank = suggest_rank(cur_rank, action)
            changed  = sug_rank != cur_rank
            dtype    = day_type(TODAY + timedelta(days=i))
            bg = C['sat'] if dtype == '土/連休初日' else C['white']

            col = i+2
            c1 = ws.cell(row=r_cur, column=col, value=cur_rank)
            c1.font = font(10, True, C['dark']); c1.fill = fill(bg)
            c1.alignment = center(); c1.border = border()

            c2 = ws.cell(row=r_sug, column=col, value=sug_rank)
            if changed and action == 'UP':
                c2.font = font(10, True, C['blue']); c2.fill = fill(C['up_bg'])
            elif changed and action == 'DOWN':
                c2.font = font(10, True, C['red']); c2.fill = fill(C['dn_bg'])
            else:
                c2.font = font(10, False, argb('#BBBBBB')); c2.fill = fill(bg)
            c2.alignment = center(); c2.border = border()

        for col_idx in range(1, NCOLS+1):
            ws.cell(row=r_spc, column=col_idx).fill = fill(C['lgray'])

    ws.freeze_panes = 'C5'

# ============================================================
# シート3: 実績稼働率ヒートマップ
# ============================================================
def build_occ_sheet(wb, daily):
    ws = wb.create_sheet("実績稼働率")
    ws.sheet_view.showGridLines = False

    title_row(ws, 1,
        f"実績稼働率ヒートマップ（2024/04〜2026/03）　総室数={TOTAL_ROOMS}室",
        14)
    note_row(ws, 2, "実予約データより算出。命名個室10室＋禁煙客室7室=17室。2026/01はデータ欠損。色: 赤=低、黄=中、緑=高", 14)

    # 月リスト・日リスト
    months = []
    d = datetime(2024, 4, 1)
    while d < datetime(2026, 4, 1):
        ym = d.strftime('%Y/%m')
        if ym not in months: months.append(ym)
        d += timedelta(days=32)
        d = d.replace(day=1)

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16
    ws.column_dimensions['A'].width = 10

    # 列ヘッダー（月）
    for mi, ym in enumerate(months, 2):
        c = ws.cell(row=3, column=mi, value=ym)
        c.font = font(9, True, C['white']); c.fill = fill(C['dark'])
        c.alignment = center(); c.border = border()
        ws.column_dimensions[get_column_letter(mi)].width = 9

    # 行ヘッダー（日付1〜31）
    for day_n in range(1, 32):
        row_n = day_n + 4
        ws.row_dimensions[row_n].height = 16
        c = ws.cell(row=row_n, column=1, value=f'{day_n}日')
        c.font = font(9, True, C['white']); c.fill = fill(C['blue'])
        c.alignment = center(); c.border = border()

    # データセル
    for mi, ym in enumerate(months, 2):
        year, month = int(ym[:4]), int(ym[5:])
        for day_n in range(1, 32):
            row_n = day_n + 4
            try:
                d = datetime(year, month, day_n)
            except:
                # その月に存在しない日
                c = ws.cell(row=row_n, column=mi, value='')
                c.fill = fill(C['lgray']); c.border = border()
                continue

            rooms = daily.get(d, 0)
            occ   = rooms / TOTAL_ROOMS
            pct   = f'{occ:.0%}'

            c = ws.cell(row=row_n, column=mi, value=occ)
            c.number_format = '0%'
            c.alignment = center(); c.border = border()

            # ヒートマップ色
            if occ == 0:
                bg = argb('#F0F0F0'); fg = argb('#AAAAAA')
            elif occ < 0.3:
                bg = argb('#FDECEA'); fg = C['red']
            elif occ < 0.5:
                bg = argb('#FEF9E7'); fg = C['orange']
            elif occ < 0.7:
                bg = argb('#EAFAF1'); fg = C['green']
            else:
                bg = argb('#1E8449'); fg = C['white']
            c.fill = fill(bg)
            c.font = font(9, occ >= 0.7, fg)

    # 月別平均行
    avg_row = 36
    ws.row_dimensions[avg_row].height = 20
    c = ws.cell(row=avg_row, column=1, value='月平均')
    c.font = font(9, True, C['white']); c.fill = fill(C['dark'])
    c.alignment = center(); c.border = border()

    for mi, ym in enumerate(months, 2):
        # Excelの平均式
        col_letter = get_column_letter(mi)
        c = ws.cell(row=avg_row, column=mi)
        c.value = f'=AVERAGE({col_letter}5:{col_letter}35)'
        c.number_format = '0.0%'
        c.font = font(9, True, C['white']); c.fill = fill(C['blue'])
        c.alignment = center(); c.border = border()

    ws.freeze_panes = 'B5'

# ============================================================
# シート4: ブッキングカーブ分析
# ============================================================
def build_curve_sheet(wb, lead_dist):
    ws = wb.create_sheet("ブッキングカーブ分析")
    ws.sheet_view.showGridLines = False
    NCOLS = 5

    title_row(ws, 1, "ブッキングカーブ分析（実予約データより）", NCOLS)
    note_row(ws, 2, "チェックイン日から何週前に予約が入ったかの分布。累積で予約進捗を把握するために使用。", NCOLS)

    day_types = ['土/連休', '日曜', '平日']
    colors    = [C['orange'], C['blue'], C['green']]

    # ヘッダー
    ws.row_dimensions[3].height = 22
    for ci, h in enumerate(['リードタイム',''+day_types[0]+'\n件数','累積%',
                             day_types[1]+'\n件数','累積%',
                             day_types[2]+'\n件数','累積%'], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = font(9, True, C['white']); c.fill = fill(C['dark'])
        c.alignment = center(); c.border = border()
    ws.column_dimensions['A'].width = 14

    NCOLS2 = 7
    for ci in range(2, NCOLS2+1):
        ws.column_dimensions[get_column_letter(ci)].width = 11

    # データ（0〜26週前）
    totals = {dt: sum(lead_dist[dt].values()) or 1 for dt in day_types}
    row = 4
    for w in range(27):
        ws.row_dimensions[row].height = 18
        label = f'{w}週前' if w > 0 else '当日〜6日前'
        c = ws.cell(row=row, column=1, value=label)
        c.font = font(10, w==0); c.fill = fill(C['gray'] if row%2==0 else C['white'])
        c.alignment = center(); c.border = border()

        col = 2
        for dt in day_types:
            cnt  = lead_dist[dt].get(w, 0)
            cum  = sum(lead_dist[dt].get(ww,0) for ww in range(w+1))
            cum_r = cum / totals[dt]

            # 件数
            c1 = ws.cell(row=row, column=col, value=cnt)
            c1.font = font(10); c1.alignment = center(); c1.border = border()
            c1.fill = fill(C['gray'] if row%2==0 else C['white'])
            # 累積%
            c2 = ws.cell(row=row, column=col+1, value=cum_r)
            c2.number_format = '0.0%'
            c2.alignment = center(); c2.border = border()
            # 累積50/75/90%をハイライト
            if cum_r >= 0.90:   bg = C['ok_bg']
            elif cum_r >= 0.75: bg = C['up_bg']
            elif cum_r >= 0.50: bg = C['sat']
            else:               bg = C['white']
            c2.fill = fill(bg)
            c2.font = font(10, cum_r >= 0.75)
            col += 2

        row += 1

    # 凡例
    row += 1
    note_row(ws, row, "■ 緑=90%超 ■ 青=75%超 ■ 黄=50%超", NCOLS2, 16)

    ws.freeze_panes = 'A4'

# ============================================================
# シート5: 競合価格モニター（実データ）
# ============================================================
def build_comp_sheet(wb, comp_prices):
    ws = wb.create_sheet("競合価格モニター")
    ws.sheet_view.showGridLines = False
    NCOLS = 1 + DAYS_AHEAD

    title_row(ws, 1, f"競合価格モニター（実データ：楽天スクレイピング）　取得日: {TODAY.strftime('%Y/%m/%d')}", NCOLS)
    note_row(ws, 2, "2食付き最安値。× = 楽天に2食付きプランなし。自社・糸柳・花水晶は楽天に2食付き掲載なしのため空欄。", NCOLS)

    # 施設順（表示順固定）
    disp_hotels = list(COMP_NAME_MAP.values())  # マッピング定義順

    ws.row_dimensions[3].height = 28
    ws.column_dimensions['A'].width = 22

    c = ws.cell(row=3, column=1, value='施設名')
    c.font = font(9, True, C['white']); c.fill = fill(C['dark'])
    c.alignment = center(); c.border = border()

    for i in range(1, DAYS_AHEAD+1):
        d = TODAY + timedelta(days=i)
        dtype = day_type(d)
        is_sat = dtype == '土/連休初日'
        bg = C['orange'] if is_sat else C['blue']
        col = i+1
        c = ws.cell(row=3, column=col, value=f"{d.month}/{d.day}\n({WDAYS[d.weekday()]})")
        c.font = font(8, True, C['white']); c.fill = fill(bg)
        c.alignment = center(); c.border = border()
        ws.column_dimensions[get_column_letter(col)].width = 7.5

    no_data_hotels = {'ホテル甲子園（自社）', '糸柳', '花水晶'}

    for hi, hotel in enumerate(disp_hotels):
        row = hi + 4
        ws.row_dimensions[row].height = 18
        is_own = hotel == 'ホテル甲子園（自社）'
        is_no_data = hotel in no_data_hotels

        bg_name = C['sub'] if is_own else (C['gray'] if hi%2==0 else C['white'])
        c = ws.cell(row=row, column=1, value=hotel)
        c.font = font(10, is_own, C['dark']); c.fill = fill(bg_name)
        c.alignment = left(); c.border = border()

        for i in range(1, DAYS_AHEAD+1):
            d = TODAY + timedelta(days=i)
            date_str = d.strftime('%Y/%m/%d')
            col = i+1
            dtype = day_type(d)
            row_bg = C['sat'] if dtype == '土/連休初日' else (C['gray'] if hi%2==0 else C['white'])

            if is_own:
                # 自社は現在ランク価格（モック）
                val = RANK_PRICE.get('F1', 33800)
                c = ws.cell(row=row, column=col, value=val)
                c.number_format = '#,##0'
                c.font = font(9, True, C['dark']); c.fill = fill(C['sub'])
            elif is_no_data:
                c = ws.cell(row=row, column=col, value='−')
                c.font = font(9, False, argb('#AAAAAA')); c.fill = fill(argb('#F5F5F5'))
            else:
                price = comp_prices.get(date_str, {}).get(hotel)
                if price is not None:
                    c = ws.cell(row=row, column=col, value=price)
                    c.number_format = '#,##0'
                    # 価格帯で色分け
                    if price >= 70000:
                        c.font = font(9, True, C['red'])
                    elif price >= 50000:
                        c.font = font(9, True, C['orange'])
                    else:
                        c.font = font(9)
                    c.fill = fill(C['sat'] if dtype == '土/連休初日' else (C['gray'] if hi%2==0 else C['white']))
                else:
                    c = ws.cell(row=row, column=col, value='×')
                    c.font = font(9, False, argb('#BBBBBB')); c.fill = fill(row_bg)
            c.alignment = center(); c.border = border()

    # 競合平均行（自社・楽天掲載なし施設を除く）
    comp_only = [h for h in disp_hotels if h not in no_data_hotels]
    avg_row = len(disp_hotels) + 4
    ws.row_dimensions[avg_row].height = 20
    c = ws.cell(row=avg_row, column=1, value=f'競合平均（{len(comp_only)}施設・価格あり日のみ）')
    c.font = font(9, True, C['white']); c.fill = fill(C['blue'])
    c.alignment = center(); c.border = border()

    for i in range(1, DAYS_AHEAD+1):
        d = TODAY + timedelta(days=i)
        date_str = d.strftime('%Y/%m/%d')
        prices = [comp_prices.get(date_str, {}).get(h) for h in comp_only]
        prices = [p for p in prices if p is not None]
        col = i+1
        c = ws.cell(row=avg_row, column=col)
        if prices:
            c.value = round(sum(prices)/len(prices))
            c.number_format = '#,##0'
            c.font = font(9, True, C['white'])
        else:
            c.value = '−'
            c.font = font(9, False, argb('#DDDDDD'))
        c.fill = fill(C['blue']); c.alignment = center(); c.border = border()

    ws.freeze_panes = 'B4'

# ============================================================
# シート6: ランク一覧（部屋×人数別価格表）
# 参照: 【ホテル甲子園様】RMシート_20240604 (1).xlsx「ランク一覧」シート
# ============================================================

# 部屋タイプ定義（定員・プレミアム・人数カット）
RANK_TABLE_ROOMS = [
    # (部屋名, 定員表示, 2名カット, 3名カット, 4名カット)
    ('スタンダード', '2〜5名',    0,  -1000,  -1000),
    ('禁煙',         '1〜5名',    0,  -1000,  -1000),
    ('昇仙峡',       '2〜4名',    0,  -1000,  -2000),
    ('本栖',         '2〜7名',    0,  -1000,  -2000),
    ('夜叉神',       '2〜6名',    0,  -1000,  -2000),
    ('天神',         '2名',       0,  None,   None ),  # 2名専用
    ('天目',         '2名',       0,  None,   None ),  # 2名専用
    ('みさか',       '2〜5名',    0,  -1000,  -1000),
    ('赤岳',         '2〜4名',    0,  -1000,  -1000),
    ('笛吹',         '2〜6名',    0,  -3000,  -3000),
    ('風林',         '2名',       0,  None,   None ),  # 高ランクで2名専用化
    ('金峰',         '2〜4名',    0,  0,      0    ),  # 人数カットなし
]

# ランク別 2名合計価格（スタンダード・禁煙は同額、他部屋は参照シートの実値）
# 各タプル: (ランク, スタンダ, 禁煙, 昇仙峡, 本栖, 夜叉神, 天神, 天目, みさか, 赤岳, 笛吹, 風林, 金峰)
RANK_TABLE_DATA = [
    #  ランク  スタンダ  禁煙   昇仙峡  本栖    夜叉神  天神   天目   みさか  赤岳   笛吹   風林   金峰
    ('A', 16300, 16300, 19800, 20300, 22800, 23500, 25800, 28800, 30800, 38300, 42000, 54000),
    ('B', 17800, 17800, 21300, 21800, 24300, 25500, 27800, 30300, 32800, 41300, 45000, 57000),
    ('C', 19300, 19300, 22800, 23300, 25800, 27500, 29800, 31800, 34800, 44300, 48000, 60000),
    ('D', 20800, 20800, 24300, 24800, 27300, 29500, 31800, 33300, 36800, 47300, 51000, 63000),
    ('E', 22300, 22300, 25800, 26300, 28800, 31500, 33800, 34800, 38800, 50300, 54000, 66000),
    ('F', 23800, 23800, 27300, 27800, 30300, 33500, 35800, 36300, 40800, 53300, 57000, 69000),
    ('I', 25300, 25300, 28800, 29300, 31800, 39500, 41800, 37800, 42800, 56300, 66000, 72000),
    ('L', 26800, 26800, 30300, 30800, 33300, 45500, 47800, 39300, 44800, 59300, 75000, 75000),
    ('O', 28300, 28300, 31800, 32300, 34800, 51500, 53800, 40800, 46800, 62300, 84000, 78000),
    ('R', 29800, 29800, 33300, 33800, 36300, 57500, 59800, 42300, 48800, 65300, 93000, 81000),
    ('U', 31300, 31300, 34800, 35300, 37800, 63500, 65800, 43800, 50800, 68300,102000, 84000),
    ('X', 32800, 32800, 36300, 36800, 39300, 69500, 71800, 45300, 52800, 71300,111000, 87000),
    ('AA',34300, 34300, 37800, 38300, 40800, 75500, 77800, 46800, 54800, 74300,120000, 90000),
    ('AD',35800, 35800, 39300, 39800, 42300, 81500, 83800, 48300, 56800, 77300,129000, 93000),
    ('AG',37300, 37300, 40800, 41300, 43800, 87500, 89800, 49800, 58800, 80300,138000, 96000),
    ('AJ',38800, 38800, 42300, 42800, 45300, 93500, 95800, 51300, 60800, 83300,147000, 99000),
    ('AM',40300, 40300, 43800, 44300, 46800, 99500,101800, 52800, 62800, 86300,156000,102000),
    ('AP',41800, 41800, 45300, 45800, 48300,     0,     0, 54300, 64800, 89300,     0,105000),
    ('AS',43300, 43300, 46800, 47300, 49800,     0,     0, 55800, 66800, 92300,     0,108000),
    ('AV',44800, 44800, 48300, 48800, 51300,     0,     0, 57300, 68800, 95300,     0,111000),
    ('AY',46300, 46300, 49800, 50300, 52800,     0,     0, 58800, 70800, 98300,     0,114000),
    ('BB',47800, 47800, 51300, 51800, 54300,     0,     0, 60300, 72800,101300,     0,117000),
]

def build_rank_table_sheet(wb):
    """部屋タイプ×ランク×人数(2〜4名)の価格表シートを生成"""
    ws = wb.create_sheet("ランク一覧")
    ws.sheet_view.showGridLines = False

    N_ROOMS = len(RANK_TABLE_ROOMS)
    PERSON_COLS = 3   # 2名 / 3名 / 4名
    NCOLS = 1 + N_ROOMS * PERSON_COLS

    room_names  = [r[0] for r in RANK_TABLE_ROOMS]
    capacities  = [r[1] for r in RANK_TABLE_ROOMS]
    cuts_3      = [r[3] for r in RANK_TABLE_ROOMS]
    cuts_4      = [r[4] for r in RANK_TABLE_ROOMS]

    # 各ランク・各部屋の人数別価格を計算
    rank_data = []
    for row in RANK_TABLE_DATA:
        rank = row[0]
        p2_list = list(row[1:])  # 部屋ごとの2名合計
        rank_data.append({
            'rank': rank,
            'p2': p2_list,
            'p3': [
                (p2 + c3) if (p2 > 0 and c3 is not None) else 0
                for p2, c3 in zip(p2_list, cuts_3)
            ],
            'p4': [
                (p2 + c4) if (p2 > 0 and c4 is not None) else 0
                for p2, c4 in zip(p2_list, cuts_4)
            ],
        })

    # ---- レイアウト ----
    N_ROOMS = len(room_names)
    PERSON_COLS = 3  # 2名/3名/4名
    NCOLS = 1 + N_ROOMS * PERSON_COLS  # ランク列 + (部屋×人数)列

    title_row(ws, 1,
        f"ランク一覧（部屋タイプ×人数別　2食付き・1名単価）　参照: {TODAY.strftime('%Y/%m/%d')}",
        NCOLS)
    note_row(ws, 2,
        "数値は1名単価（円）。2名合計は×2してください。0は対応不可（定員外または廃止）。",
        NCOLS)

    # ---- 部屋名ヘッダー（行3: 部屋名スパン） ----
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16

    # ランク列ヘッダー
    for rn in [3, 4, 5]:
        c = ws.cell(row=rn, column=1, value='ランク' if rn == 3 else '')
        c.font = font(10, True, C['white']); c.fill = fill(C['dark'])
        c.alignment = center(); c.border = border()
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
    ws.column_dimensions['A'].width = 8

    room_colors = [
        argb('#1A5276'), argb('#1A5276'), argb('#6C3483'), argb('#6C3483'),
        argb('#884EA0'), argb('#117A65'), argb('#117A65'), argb('#0E6655'),
        argb('#0E6655'), argb('#1A5276'), argb('#784212'), argb('#784212'),
    ]

    for ri, room in enumerate(room_names):
        start_col = 2 + ri * PERSON_COLS
        room_bg = room_colors[ri % len(room_colors)]

        # 部屋名（3行にスパン → 3行目:部屋名, 4行目:定員, 5行目:人数ヘッダー）
        c = ws.cell(row=3, column=start_col, value=room)
        c.font = font(9, True, C['white']); c.fill = fill(room_bg)
        c.alignment = center(); c.border = border()
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + PERSON_COLS - 1)

        cap_val = capacities[ri]
        c4 = ws.cell(row=4, column=start_col, value=f'定員:{cap_val}')
        c4.font = font(8, False, C['white']); c4.fill = fill(room_bg)
        c4.alignment = center(); c4.border = border()
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + PERSON_COLS - 1)

        for pi, (lbl, cut_info) in enumerate(zip(['2名', '3名', '4名'], [0, cuts_3[ri], cuts_4[ri]])):
            col = start_col + pi
            cut_str = f'({cut_info:+,})' if (cut_info is not None and cut_info < 0) else ''
            c5 = ws.cell(row=5, column=col, value=f'{lbl}{cut_str}')
            c5.font = font(8, True, C['white']); c5.fill = fill(room_bg)
            c5.alignment = center(); c5.border = border()
            ws.column_dimensions[get_column_letter(col)].width = 9

    # ---- ランクデータ行 ----
    # ランクごとの色（ラクツウ実際の色に合わせる）
    # フォーマット: AARRGGBB（FFで不透明）
    RANK_COLORS = {
        'A':  'FFFF0000',  # 赤
        'B':  'FF00FFFF',  # シアン
        'C':  'FFFFFF00',  # 黄
        'D':  'FF0080FF',  # 青
        'E':  'FF804040',  # 茶
        'F':  'FFFF00FF',  # マゼンタ
        'I':  'FF808000',  # オリーブ
        'L':  'FF400040',  # 濃紫
        'O':  'FF0000A0',  # 濃青
        'R':  'FF008040',  # 緑
        'U':  'FFF5390F',  # オレンジ赤
        'X':  'FFA6FF00',  # 黄緑
        'AA': 'FFBB00FF',  # 紫
        'AD': 'FFCC0000',  # 濃赤
        'AG': 'FF469154',  # 緑
        'AJ': 'FFB591B5',  # ラベンダー
        'AM': 'FF8BC28A',  # 薄緑
        'AP': 'FFE8CAE8',  # 薄ピンク
        'AS': 'FFA3A331',  # カーキ
        'AV': 'FFF5877F',  # サーモン
        'AY': 'FFB06A2C',  # 茶
        'BB': 'FF4D0D29',  # 濃赤茶
    }

    for di, rd in enumerate(rank_data):
        row_n = 6 + di
        ws.row_dimensions[row_n].height = 18
        rank_bg = RANK_COLORS.get(rd['rank'], 'FFFFFFFF')
        # 文字色：背景が暗い場合は白、明るい場合は黒
        r_val = int(rank_bg[2:4], 16)
        g_val = int(rank_bg[4:6], 16)
        b_val = int(rank_bg[6:8], 16)
        lum = 0.299 * r_val + 0.587 * g_val + 0.114 * b_val
        rank_fg = C['white'] if lum < 140 else C['dark']
        row_bg = rank_bg  # データ行は薄い色にする（ランク色の10%）

        # ランク名セル（ラクツウ色）
        c = ws.cell(row=row_n, column=1, value=rd['rank'])
        c.font = font(11, True, rank_fg); c.fill = fill(rank_bg)
        c.alignment = center(); c.border = border()

        for ri in range(N_ROOMS):
            start_col = 2 + ri * PERSON_COLS
            room_bg = room_colors[ri % len(room_colors)]

            for pi, prices in enumerate([rd['p2'], rd['p3'], rd['p4']]):
                col = start_col + pi
                price = prices[ri]
                c = ws.cell(row=row_n, column=col, value=price if price > 0 else None)
                c.alignment = center(); c.border = border()

                if price == 0:
                    c.value = '−'
                    c.font = font(9, False, argb('#CCCCCC'))
                    c.fill = fill(argb('#F5F5F5'))
                else:
                    c.number_format = '#,##0'
                    # 2名列はやや濃い背景、3名/4名は白
                    if pi == 0:
                        # ランク色を薄めた背景（RGB各チャンネルを白(255)と blend 30%）
                        br = min(255, r_val + int((255 - r_val) * 0.7))
                        bg_val = min(255, g_val + int((255 - g_val) * 0.7))
                        bb_val = min(255, b_val + int((255 - b_val) * 0.7))
                        cell_bg = f'FF{br:02X}{bg_val:02X}{bb_val:02X}'
                        c.font = font(10, True, C['dark'])
                        c.fill = fill(cell_bg)
                    else:
                        c.font = font(9, False, argb('#444444'))
                        c.fill = fill(C['white'] if di % 2 == 0 else C['altrow'])

    # ---- 凡例行 ----
    legend_row = 6 + len(rank_data) + 1
    ws.row_dimensions[legend_row].height = 18
    c = ws.cell(row=legend_row, column=1,
        value='※ 一朝食は各価格から5,000円引き　素泊りは8,000円引き　エステ付きは7,200円増し')
    c.font = font(9, False, argb('#888888')); c.fill = fill(C['gray'])
    c.alignment = left()
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=NCOLS)

    ws.freeze_panes = 'B6'

# ============================================================
# シート7: 月次売上サマリー
# ============================================================
def build_monthly_summary_sheet(wb, monthly_rev):
    """
    月次売上サマリー（前年〜来期）
    - 予算 / 現況売上 / 差異 / 達成率% を月別に一覧
    - 過去月=実績（確定）、当月〜未来=未確定（現時点予約ベース）
    - 前年同月比を右端に表示
    """
    ws = wb.create_sheet("月次売上サマリー")
    ws.sheet_view.showGridLines = False

    # 表示月範囲：13ヶ月前〜6ヶ月後
    cur_ym = TODAY.replace(day=1)
    start_ym = (cur_ym.replace(month=cur_ym.month - 1) if cur_ym.month > 1
                else cur_ym.replace(year=cur_ym.year - 1, month=12))
    # 13ヶ月前から開始
    months = []
    d = cur_ym
    for _ in range(13):
        d = (d.replace(month=d.month - 1) if d.month > 1
             else d.replace(year=d.year - 1, month=12))
    start = d
    d = start
    while d <= cur_ym.replace(month=min(cur_ym.month + 6, 12),
                               year=cur_ym.year + (1 if cur_ym.month + 6 > 12 else 0)):
        months.append(d.strftime('%Y%m'))
        if d.month == 12:
            d = d.replace(year=d.year + 1, month=1)
        else:
            d = d.replace(month=d.month + 1)

    NCOLS = 1 + len(months) + 1  # ラベル列 + 月列 + 前年比列

    title_row(ws, 1,
        f"月次売上サマリー（実績・現況）　集計日: {TODAY.strftime('%Y/%m/%d')}",
        NCOLS)
    note_row(ws, 2,
        "売上=全科目合算（宿泊+食事+追加）。予算は config の MONTHLY_BUDGET を編集してください。未確定=当月以降の現時点予約ベース。",
        NCOLS)

    # ---- 列ヘッダー ----
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18

    # ラベル列
    ws.column_dimensions['A'].width = 16
    for rn, lbl in [(3, ''), (4, '')]:
        c = ws.cell(row=rn, column=1, value=lbl)
        c.fill = fill(C['dark']); c.border = border(); c.alignment = center()

    # 月列
    cur_month_str = cur_ym.strftime('%Y%m')
    for ci, ym in enumerate(months, 2):
        y, m = int(ym[:4]), int(ym[4:])
        label = f"{y}年{m}月"
        is_future = ym >= cur_month_str
        is_ly = ym < (cur_ym.replace(year=cur_ym.year - 1)).strftime('%Y%m')  # 前年以前

        if is_future:
            hdr_bg = argb('#E67E22')   # オレンジ = 未確定
            status = '未確定'
        else:
            hdr_bg = C['dark']         # ネイビー = 実績
            status = '実績'

        # 年月行
        c3 = ws.cell(row=3, column=ci, value=label)
        c3.font = font(9, True, C['white']); c3.fill = fill(hdr_bg)
        c3.alignment = center(); c3.border = border()
        # ステータス行
        c4 = ws.cell(row=4, column=ci, value=status)
        c4.font = font(8, False, C['white']); c4.fill = fill(hdr_bg)
        c4.alignment = center(); c4.border = border()

        ws.column_dimensions[get_column_letter(ci)].width = 14

    # 昨年実績対比列（右端）
    last_col = len(months) + 2
    ws.column_dimensions[get_column_letter(last_col)].width = 14
    c3 = ws.cell(row=3, column=last_col, value='昨年\n実績対比')
    c3.font = font(9, True, C['white']); c3.fill = fill(C['blue'])
    c3.alignment = center(); c3.border = border()
    c4 = ws.cell(row=4, column=last_col, value='今年/昨年')
    c4.font = font(8, False, C['white']); c4.fill = fill(C['blue'])
    c4.alignment = center(); c4.border = border()

    # ---- データ行 ----
    ROW_DEFS = [
        ('予算（目標）',     C['lgray'],  argb('#444444'), '#,##0'),
        ('現況売上',         C['white'],  C['dark'],       '#,##0'),
        ('差異（現況-予算）', argb('#FFF8F0'), argb('#444444'), '+#,##0;-#,##0;"-"'),
        ('達成率',           C['white'],  C['dark'],       '0.0%'),
    ]

    for ri, (label, row_bg, row_fg, fmt) in enumerate(ROW_DEFS):
        data_row = 5 + ri
        ws.row_dimensions[data_row].height = 22

        # ラベルセル
        c = ws.cell(row=data_row, column=1, value=label)
        c.font = font(10, ri in (0, 3), C['dark']); c.fill = fill(C['lgray'])
        c.alignment = left(); c.border = border()

        # 月別セル
        total_actual = 0.0
        total_budget = 0.0
        total_ly_actual = 0.0

        for ci, ym in enumerate(months, 2):
            budget = MONTHLY_BUDGET.get(ym)
            actual = monthly_rev.get(ym, 0.0)
            is_future = ym >= cur_month_str

            # 前年同月
            ly_ym = f'{int(ym[:4])-1}{ym[4:]}'
            ly_actual = monthly_rev.get(ly_ym, 0.0)

            # 集計用（昨年対比計算のため今年分のみ）
            y = int(ym[:4])
            if y == TODAY.year:
                total_actual  += actual
                total_budget  += (budget or 0)
                # 前年同月実績
                total_ly_actual += ly_actual

            # セル値
            if ri == 0:    # 予算
                val = budget
            elif ri == 1:  # 現況売上
                val = actual if actual > 0 else (None if is_future else 0.0)
            elif ri == 2:  # 差異
                val = (actual - budget) if (budget and actual) else None
            else:          # 達成率
                val = (actual / budget) if (budget and actual > 0) else None

            c = ws.cell(row=data_row, column=ci, value=val)
            c.number_format = fmt
            c.alignment = center(); c.border = border()

            # 背景色
            if ri == 3 and val is not None:  # 達成率
                if val >= 1.0:    bg = argb('#D5F5E3'); c.font = font(10, True, C['green'])
                elif val >= 0.8:  bg = argb('#FDEBD0'); c.font = font(10, False, C['orange'])
                else:             bg = argb('#FADBD8'); c.font = font(10, True, C['red'])
                c.fill = fill(bg)
            elif ri == 2 and val is not None:  # 差異
                if val >= 0: c.font = font(10, False, C['green'])
                else:        c.font = font(10, False, C['red'])
                c.fill = fill(C['white'] if not is_future else argb('#FFF8F0'))
            else:
                c.font = font(10, ri == 0, row_fg)
                if is_future:
                    c.fill = fill(argb('#FFF8F0'))  # 未確定は薄いオレンジ背景
                else:
                    c.fill = fill(row_bg)

        # 昨年実績対比セル（今年合計 / 前年合計）
        c_ly = ws.cell(row=data_row, column=last_col)
        c_ly.alignment = center(); c_ly.border = border()
        if ri == 1:  # 現況売上行にのみ前年比を表示
            if total_ly_actual > 0:
                yoy = total_actual / total_ly_actual
                c_ly.value = yoy
                c_ly.number_format = '0.0%'
                if yoy >= 1.0:
                    c_ly.font = font(11, True, C['green']); c_ly.fill = fill(argb('#D5F5E3'))
                elif yoy >= 0.9:
                    c_ly.font = font(11, True, C['orange']); c_ly.fill = fill(argb('#FDEBD0'))
                else:
                    c_ly.font = font(11, True, C['red']); c_ly.fill = fill(argb('#FADBD8'))
            else:
                c_ly.value = '−'; c_ly.font = font(10, False, argb('#AAAAAA'))
                c_ly.fill = fill(C['lgray'])
        else:
            c_ly.value = ''; c_ly.fill = fill(C['lgray'])

    # 罫線区切り（行グループの下に太線）
    thick = Side(style='medium', color='FF999999')
    for ci in range(1, NCOLS + 1):
        c = ws.cell(row=8, column=ci)
        c.border = Border(bottom=thick, left=side(), right=side(), top=side())

    # メモ行
    memo_row = 10
    ws.row_dimensions[memo_row].height = 18
    c = ws.cell(row=memo_row, column=1,
        value='※ 現況売上 = PMSデータの金額合計（宿泊代・食事・追加料金すべて含む）。子供料金(宿泊子Ａ等)も含む。')
    c.font = font(8, False, argb('#888888')); c.fill = fill(C['gray'])
    c.alignment = left()
    ws.merge_cells(start_row=memo_row, start_column=1, end_row=memo_row, end_column=NCOLS)

    ws.freeze_panes = 'B5'

# ============================================================
# シート7: 凡例・使い方
# ============================================================
def build_legend_sheet(wb):
    ws = wb.create_sheet("凡例・使い方")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 25

    title_row(ws, 1, "石和温泉 ホテル甲子園　RMシステム　凡例・使い方", 4)

    def sec(r, text):
        ws.row_dimensions[r].height = 22
        c = ws.cell(row=r, column=2, value=text)
        c.font = font(11, True, C['white']); c.fill = fill(C['dark'])
        c.alignment = left()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    def row2(r, b, c_val, bg_b=None, fg_b=None):
        ws.row_dimensions[r].height = 20
        cb = ws.cell(row=r, column=2, value=b)
        cb.font = font(10, bool(fg_b), fg_b or C['dark'])
        cb.fill = fill(bg_b or C['gray']); cb.alignment = center(); cb.border = border()
        cc = ws.cell(row=r, column=3, value=c_val)
        cc.font = font(10); cc.fill = fill(C['white']); cc.alignment = left(); cc.border = border()
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    r = 3
    sec(r, '【シート構成】'); r+=1
    for nm, desc in [
        ('RM価格案',         '日別30日間の推奨アクション（STAY/UP/DOWN）と推奨ランク'),
        ('部屋別ランク設定', '部屋タイプ×日付マトリクス。現在ランクと推奨ランクを一覧'),
        ('実績稼働率',       '2年分の実予約データによる日別稼働率ヒートマップ（実データ）'),
        ('ブッキングカーブ分析', '曜日タイプ別の予約リードタイム分布（実データ）'),
        ('競合価格モニター', '競合9施設の価格を30日分で比較（現在はモック）'),
        ('ランク一覧',       '部屋タイプ×ランク×人数(2〜4名)の1名単価一覧（実データ）'),
        ('月次売上サマリー', '月別予算 vs 実績/現況・達成率・前年比（実データ）'),
        ('凡例・使い方',     'この画面'),
    ]:
        row2(r, nm, desc, C['sub'], C['dark']); r+=1

    r+=1
    sec(r, '【推奨アクション】'); r+=1
    for lbl, bg, fg, desc in [
        ('UP検討',   C['up_bg'], C['blue'],  '消化率が目標を+5%以上上回る → ランクを1段上げる'),
        ('STAY',     C['ok_bg'], C['green'], '消化率が目標の±5〜10%以内 → 現在ランクを維持'),
        ('DOWN推奨', C['dn_bg'], C['red'],   '消化率が目標を-10%以上下回る → ランクを1段下げて集客強化'),
    ]:
        row2(r, lbl, desc, bg, fg); r+=1

    r+=1
    sec(r, '【ランク体系】'); r+=1
    ws.row_dimensions[r].height = 18
    for ci, h in enumerate(['ランク','1名単価','2名合計（2食付き）'], 2):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = font(9, True, C['white']); c.fill = fill(C['blue'])
        c.alignment = center(); c.border = border()
    r+=1
    for rank in RANKS:
        ws.row_dimensions[r].height = 16
        ws.row_dimensions[r].height = 16
        cb = ws.cell(row=r, column=2, value=rank)
        cb.font = font(10, False, C['dark']); cb.fill = fill(C['gray'] if r%2==0 else C['white'])
        cb.alignment = center(); cb.border = border()
        # 1名単価
        c1 = ws.cell(row=r, column=3, value=f'¥{RANK_PRICE[rank]:,}')
        c1.font = font(10); c1.fill = fill(C['gray'] if r%2==0 else C['white'])
        c1.alignment = center(); c1.border = border()
        # 2名合計
        c2 = ws.cell(row=r, column=4, value=f'¥{RANK_PRICE[rank]*2:,}')
        c2.font = font(10, True, C['dark']); c2.fill = fill(C['gray'] if r%2==0 else C['white'])
        c2.alignment = center(); c2.border = border()
        r+=1

    r+=1
    sec(r, '【注意事項】'); r+=1
    notes = [
        f'総室数={TOTAL_ROOMS}室（命名個室10室＋禁煙客室7室）。変更が必要な場合は export_rm_excel.py の TOTAL_ROOMS を修正してください',
        '競合価格はモック。competitor_scraper.py で取得した実データへ差し替え予定',
        'ランク変更はラクツウのサイトコントローラー画面から手動で反映してください',
        '消化率「実績（前年同週）」は前年の同じ曜日週の実績を参照しています',
    ]
    for note in notes:
        ws.row_dimensions[r].height = 20
        c = ws.cell(row=r, column=2, value=f'・{note}')
        c.font = font(9, False, C['red']); c.fill = fill(argb('#FFFBF0'))
        c.alignment = left()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r+=1

# ============================================================
# メイン
# ============================================================
def run():
    print("データ読み込み中...")
    daily, lead_dist, comp_prices, data_source, monthly_rev = load_data()
    print(f"  予約データ読込完了: {data_source}（総室数={TOTAL_ROOMS}室）")
    comp_dates = len(comp_prices)
    print(f"  競合価格データ: {comp_dates}日分読込完了")
    print(f"  月別売上集計: {len(monthly_rev)}ヶ月分")

    print("Excelを生成中...")
    wb = Workbook()
    wb.remove(wb.active)

    ws1, rows_data = build_rm_sheet(wb, daily, comp_prices, data_source)
    print("  ① RM価格案 完了")
    build_rank_sheet(wb, rows_data)
    print("  ② 部屋別ランク設定 完了")
    build_occ_sheet(wb, daily)
    print("  ③ 実績稼働率ヒートマップ 完了")
    build_curve_sheet(wb, lead_dist)
    print("  ④ ブッキングカーブ分析 完了")
    build_comp_sheet(wb, comp_prices)
    print("  ⑤ 競合価格モニター 完了")
    build_rank_table_sheet(wb)
    print("  ⑥ ランク一覧 完了")
    build_monthly_summary_sheet(wb, monthly_rev)
    print("  ⑦ 月次売上サマリー 完了")
    build_legend_sheet(wb)
    print("  ⑧ 凡例・使い方 完了")

    wb.save(OUTPUT_PATH)
    print(f"\n完了！ → {OUTPUT_PATH}")

if __name__ == "__main__":
    run()
